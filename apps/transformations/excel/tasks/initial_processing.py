import hashlib
import os
import tempfile
from datetime import timedelta
from pathlib import Path

from openpyxl import load_workbook
from prefect import task

from packages.filestorage import filestorage
from packages.schemas.satu import UserFilledData


@task(
    name="Extract Product Info from Excel",
    result_serializer="json",
    cache_expiration=timedelta(days=1),
    tags=["excel"],
)
async def initial_excel_processing(id: str) -> list[dict]:
    """
    Extract Product Info from Excel

    :param id: The ID of the Excel file
    :return: The extracted info
    """

    file = filestorage.download(id)

    with tempfile.TemporaryDirectory() as tmpdirname:
        excel = Path(tmpdirname) / f"{os.urandom(16).hex()}.xlsx"
        excel.write_bytes(file)
        workbook = load_workbook(excel)
        if not (ws := workbook.active):
            raise ValueError("No active worksheet found")

        header = [str(cell.value) for cell in ws[1]]
        data = [
            {
                str(header[col]): str(cell.value)
                for col, cell in enumerate(row)
                if cell.value
            }
            for row in ws.iter_rows(min_row=2)
        ]

    all_cols = set(header)
    primary_cols = {col for col in all_cols if all(col in row for row in data)}

    if not primary_cols:
        raise ValueError("No primary columns found")

    return [
        {
            "hash": hashlib.sha256(row["sku"].encode()).hexdigest(),
            "url": "",
            "data": {
                prop: row[prop]
                for prop in UserFilledData.model_fields.keys()
                if prop in row
            },
            "source_id": id,
            "reprocessing": True,
        }
        for row in data
    ]
